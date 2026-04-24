import json
import os
import base64
import csv
import io
import mimetypes
from urllib.parse import quote
from datetime import date
from functools import partial
from http import cookies
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse

from auth import (
    authenticate_user,
    create_session,
    create_student_user,
    create_teacher_user,
    delete_teacher_user,
    delete_session,
    delete_student_user,
    get_teacher_managed_classes,
    get_user_by_session,
    init_database,
    list_school_users,
    list_students_for_classes,
)
from planner import chat_reply, extract_teacher_homework, generate_analysis_reply, generate_note_subtasks, smart_processor
from user_store import (
    archive_overdue_notes,
    build_points_profile,
    clear_archive_trash,
    complete_note,
    create_teacher_homework_assignment,
    delete_teacher_homework_assignment,
    ensure_user_storage,
    get_student_homework_submission,
    get_chat_session,
    list_chat_sessions,
    load_completed_tasks,
    load_state,
    load_student_grades,
    load_student_performance,
    move_archive_item_to_trash,
    save_note_subtasks,
    save_state,
    save_chat_exchange,
    toggle_note_subtask,
    build_student_ai_notifications,
    submit_student_homework,
    save_planner_result,
    sync_student_homework_rating,
    toggle_student_homework_completion,
)


SESSION_COOKIE_NAME = "jarvis_session"
DB_FILENAME = "jarvis_study.db"


ADMIN_IMPORT_HEADERS = {
    "имя ученика": "display_name",
    "ученик": "display_name",
    "display_name": "display_name",
    "name": "display_name",
    "класс": "class_name",
    "class": "class_name",
    "class_name": "class_name",
    "логин": "username",
    "username": "username",
    "пароль": "password",
    "password": "password",
}


def _split_managed_classes(raw_value: str | None) -> list[str]:
    return [part.strip() for part in str(raw_value or "").split(",") if part.strip()]


def _coerce_numeric_grade(value: str | int | float | None) -> float | None:
    try:
        numeric = float(str(value or "").strip().replace(",", "."))
    except (TypeError, ValueError):
        return None
    return numeric if numeric > 0 else None


def _build_analysis_context(data_root: str, storage_key: str, parsed: dict[str, object]) -> dict[str, object]:
    scope = [
        str(item).strip().lower()
        for item in (parsed.get("scope") or [])
        if str(item).strip().lower() in {"notes", "homework", "grades", "performance"}
    ]
    if not scope:
        scope = ["notes", "homework", "grades", "performance"]

    state = load_state(data_root, storage_key)
    context: dict[str, object] = {
        "scope": scope,
        "plan_date": parsed.get("plan_date"),
        "focus": parsed.get("focus"),
    }

    if "notes" in scope:
        notes: list[dict[str, object]] = []
        for item in state.get("notes", [])[:12]:
            result = item.get("result") if isinstance(item.get("result"), dict) else {}
            notes.append(
                {
                    "task": result.get("task") or item.get("source_text"),
                    "date": result.get("date"),
                    "saved_at": item.get("saved_at"),
                }
            )
        context["notes"] = notes

    if "homework" in scope:
        homework_items: list[dict[str, object]] = []
        for item in state.get("homework", [])[:20]:
            homework_items.append(
                {
                    "subject": item.get("subject"),
                    "task": item.get("task"),
                    "date": item.get("date"),
                    "done": bool(item.get("done", False)),
                    "class_name": item.get("class_name"),
                    "priority": item.get("priority"),
                    "volume": item.get("volume"),
                    "points_value": item.get("points_value"),
                    "homework_grade": item.get("homework_grade"),
                }
            )
        context["homework"] = homework_items

    if "grades" in scope:
        grades_payload = load_student_grades(data_root, storage_key)
        current_quarter_id = str(grades_payload.get("current_quarter") or "").strip()
        quarters = grades_payload.get("quarters") or []
        quarter = next((item for item in quarters if item.get("id") == current_quarter_id), quarters[0] if quarters else None)
        subjects_summary: list[dict[str, object]] = []
        if isinstance(quarter, dict):
            for subject in quarter.get("subjects", [])[:12]:
                subjects_summary.append(
                    {
                        "name": subject.get("name"),
                        "average": subject.get("average"),
                        "grades": list((subject.get("grades") or [])[:8]),
                    }
                )
        context["grades"] = {
            "quarter": (quarter or {}).get("label") if isinstance(quarter, dict) else None,
            "points": grades_payload.get("points"),
            "points_profile": grades_payload.get("points_profile"),
            "streak": grades_payload.get("streak"),
            "weekly_goal": grades_payload.get("weekly_goal"),
            "assistant_subject": grades_payload.get("assistant_subject"),
            "subjects": subjects_summary,
        }

    if "performance" in scope:
        performance_payload = load_student_performance(data_root, storage_key)
        days_summary: list[dict[str, object]] = []
        for day in (performance_payload.get("days") or [])[:3]:
            lessons_summary: list[dict[str, object]] = []
            for lesson in (day.get("lessons") or [])[:8]:
                lessons_summary.append(
                    {
                        "number": lesson.get("number"),
                        "subject": lesson.get("subject"),
                        "time": lesson.get("time"),
                        "task": lesson.get("task"),
                        "status": lesson.get("status"),
                    }
                )
            days_summary.append(
                {
                    "date": day.get("date"),
                    "title": day.get("title"),
                    "lessons": lessons_summary,
                }
            )
        context["performance"] = {"days": days_summary}

    return context


def _build_teacher_grades_payload(db_path: str, data_root: str, teacher_user: dict[str, object], requested_quarter_id: str | None = None) -> dict[str, object]:
    students = list_students_for_classes(db_path, ["10 А", "10 Б", "11 А"])
    section_specs = [
        {
            "title": "Алгебра, 10 А",
            "subject": "Алгебра",
            "student_usernames": ["student", "student4", "student5"],
        },
        {
            "title": "Алгебра, 10 Б",
            "subject": "Алгебра",
            "student_usernames": ["student2"],
        },
        {
            "title": "Алгебра, 11 А",
            "subject": "Алгебра",
            "student_usernames": ["student3"],
        },
    ]

    sections: list[dict[str, object]] = []
    current_quarter_id = "q4"
    quarter_label = "4 четверть"

    for section_spec in section_specs:
        class_students = [
            item for item in students
            if str(item.get("username") or "").strip() in section_spec["student_usernames"]
        ]
        if not class_students:
            continue

        rows: list[dict[str, object]] = []
        columns: list[str] = []
        completion_values: list[float] = []
        risk_students: list[dict[str, object]] = []

        for student in class_students:
            storage_key = str(student.get("storage_key") or "").strip()
            if not storage_key:
                continue

            ensure_user_storage(data_root, storage_key)
            grades_payload = load_student_grades(data_root, storage_key)
            quarter_id = str(requested_quarter_id or grades_payload.get("current_quarter") or current_quarter_id).strip() or current_quarter_id
            quarters = grades_payload.get("quarters") or []
            quarter = next((item for item in quarters if item.get("id") == quarter_id), quarters[0] if quarters else None)
            if not quarter:
                continue

            current_quarter_id = str(quarter.get("id") or current_quarter_id)
            quarter_label = str(quarter.get("label") or quarter_label)
            if not columns:
                columns = list((quarter.get("columns") or [])[:6])

            completion_value = _coerce_numeric_grade(quarter.get("completion_percent"))
            if completion_value is not None:
                completion_values.append(completion_value)

            subject_name = str(section_spec.get("subject") or "").strip()
            subjects = quarter.get("subjects") or []
            target_subject = next((item for item in subjects if str(item.get("name") or "").strip() == subject_name), None)
            if not target_subject:
                continue

            average_label = str(target_subject.get("average") or "").strip()
            average_numeric = _coerce_numeric_grade(average_label)
            tone = (
                "good"
                if (average_numeric is not None and average_numeric >= 4.0)
                else "warn"
                if (average_numeric is not None and average_numeric == 3.0)
                else "bad"
                if (average_numeric is not None and average_numeric <= 2.0)
                else "warn"
            )
            rows.append(
                {
                    "name": str(student.get("display_name") or student.get("username") or "Ученик"),
                    "grades": list((target_subject.get("grades") or [])[:len(columns)]),
                    "average": average_label,
                    "tone": tone,
                    "points_score": int((grades_payload.get("points") or {}).get("score") or 0),
                    "points_title": str((grades_payload.get("points_profile") or {}).get("title") or ""),
                }
            )
            if average_numeric is not None:
                risk_students.append(
                    {
                        "key": str(student.get("display_name") or student.get("username") or "Ученик"),
                        "value": max(0, min(5, round(average_numeric))),
                        "tone": tone,
                    }
                )

        if rows:
            sections.append(
                {
                    "title": str(section_spec["title"]),
                    "columns": columns,
                    "students": rows,
                    "completion_percent": round(sum(completion_values) / len(completion_values)) if completion_values else 0,
                    "risk_students": risk_students,
                }
            )

    return {
        "title": "Успеваемость классов, отметки за домашние задания",
        "current_quarter": current_quarter_id,
        "quarter_label": quarter_label,
        "sections": sections,
    }


def _build_teacher_archive_payload(
    db_path: str,
    data_root: str,
    teacher_user: dict[str, object],
    selected_subject: str | None = None,
    selected_date: str | None = None,
    selected_class: str | None = None,
) -> dict[str, object]:
    teacher_storage_key = str(teacher_user.get("storage_key") or "").strip()
    if not teacher_storage_key:
        return {"filters": {"subjects": [], "dates": [], "classes": []}, "selected": {}, "rows": [], "completion_percent": 0}

    teacher_state = load_state(data_root, teacher_storage_key)
    homework_items = [
        item for item in teacher_state.get("homework", [])
        if isinstance(item, dict) and str(item.get("class_name") or "").strip() and str(item.get("subject") or "").strip()
    ]
    if not homework_items:
        return {"filters": {"subjects": [], "dates": [], "classes": []}, "selected": {}, "rows": [], "completion_percent": 0}

    subjects = sorted({str(item.get("subject") or "").strip() for item in homework_items})
    dates = sorted({str(item.get("date") or "").strip() for item in homework_items})
    classes = sorted({str(item.get("class_name") or "").strip() for item in homework_items})

    selected_item = next(
        (
            item for item in homework_items
            if (not str(selected_subject or "").strip() or str(item.get("subject") or "").strip() == str(selected_subject or "").strip())
            and (not str(selected_date or "").strip() or str(item.get("date") or "").strip() == str(selected_date or "").strip())
            and (not str(selected_class or "").strip() or str(item.get("class_name") or "").strip() == str(selected_class or "").strip())
        ),
        homework_items[0],
    )
    selected_subject = str(selected_item.get("subject") or "")
    selected_date = str(selected_item.get("date") or "")
    selected_class = str(selected_item.get("class_name") or "")

    selected_students = list_students_for_classes(db_path, [selected_class])
    rows: list[dict[str, object]] = []
    completion_values: list[float] = []

    for index, student in enumerate(selected_students):
        storage_key = str(student.get("storage_key") or "").strip()
        if not storage_key:
            continue
        ensure_user_storage(data_root, storage_key)
        student_state = load_state(data_root, storage_key)
        homework_entry = next(
            (
                item for item in student_state.get("homework", [])
                if str(item.get("subject") or "").strip() == selected_subject
                and str(item.get("date") or "").strip() == selected_date
                and str(item.get("class_name") or "").strip() == selected_class
            ),
            None,
        )

        submitted_at = str((homework_entry or {}).get("submitted_at") or "").strip()
        grade_value = str((homework_entry or {}).get("homework_grade") or "").strip()
        is_done = bool(submitted_at)
        if is_done:
            completion_values.append(100)
        done_label = f"сделано {submitted_at}" if is_done else "не сделано"
        rows.append(
            {
                "name": str(student.get("display_name") or student.get("username") or "??????"),
                "status_text": done_label,
                "submitted": is_done,
                "file_name": str((homework_entry or {}).get("submitted_file_name") or "").strip(),
                "student_username": str(student.get("username") or "").strip(),
                "student_homework_id": str((homework_entry or {}).get("id") or "").strip(),
                "grade": grade_value,
                "tone": "good" if grade_value in {"4", "5"} else "warn" if grade_value == "3" else "bad" if grade_value == "2" else "none",
            }
        )

    return {
        "filters": {
            "subjects": subjects,
            "dates": dates,
            "classes": classes,
        },
        "selected": {
            "subject": selected_subject,
            "date": selected_date,
            "class_name": selected_class,
        },
        "rows": rows,
        "completion_percent": round(sum(completion_values) / len(rows)) if rows else 0,
    }


def _normalize_import_header(value: object) -> str:
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def _extract_students_from_csv(file_bytes: bytes) -> list[dict[str, str]]:
    text = None
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("Не удалось прочитать CSV-файл")

    reader = csv.DictReader(io.StringIO(text))
    return _normalize_import_rows(reader)


def _extract_students_from_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("Для Excel-импорта установите openpyxl") from error

    workbook = load_workbook(io.BytesIO(file_bytes), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [_normalize_import_header(cell) for cell in rows[0]]
    normalized_rows: list[dict[str, object]] = []
    for row in rows[1:]:
        if not any(cell not in (None, "") for cell in row):
            continue
        normalized_rows.append({headers[index]: row[index] for index in range(min(len(headers), len(row)))})
    return _normalize_import_rows(normalized_rows)


def _normalize_import_rows(rows: list[dict[str, object]] | csv.DictReader) -> list[dict[str, str]]:
    students: list[dict[str, str]] = []
    for raw_row in rows:
        mapped: dict[str, str] = {}
        for header, value in dict(raw_row).items():
            field_name = ADMIN_IMPORT_HEADERS.get(_normalize_import_header(header))
            if not field_name:
                continue
            mapped[field_name] = str(value or "").strip()
        if any(mapped.get(field) for field in ("display_name", "class_name", "username", "password")):
            students.append(mapped)
    return students


def _extract_students_from_import(file_name: str, file_bytes: bytes) -> list[dict[str, str]]:
    lowered = file_name.lower()
    if lowered.endswith(".csv"):
        return _extract_students_from_csv(file_bytes)
    if lowered.endswith(".xlsx"):
        return _extract_students_from_xlsx(file_bytes)
    raise ValueError("Поддерживаются только файлы .xlsx и .csv")


class AppHandler(SimpleHTTPRequestHandler):
    db_path: str
    data_root: str

    def _send_json(self, status_code: int, payload: dict, extra_headers: list[tuple[str, str]] | None = None) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status_code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        for header_name, header_value in extra_headers or []:
            self.send_header(header_name, header_value)
        self.end_headers()
        self.wfile.write(body)

    def _send_html(self, filename: str) -> None:
        target = os.path.join(self.directory, filename)
        try:
            with open(target, "rb") as source:
                body = source.read()
        except FileNotFoundError:
            self.send_error(404, "File not found")
            return

        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _send_file(self, file_path: str, download_name: str) -> None:
        try:
            with open(file_path, "rb") as source:
                body = source.read()
        except FileNotFoundError:
            self.send_error(404, "File not found")
            return

        content_type = mimetypes.guess_type(download_name or file_path)[0] or "application/octet-stream"
        safe_name = os.path.basename(download_name or "solution.bin")
        ascii_name = safe_name.encode("ascii", "ignore").decode("ascii").strip() or "solution.bin"
        quoted_name = quote(safe_name, safe="")
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Content-Disposition", f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quoted_name}")
        self.end_headers()
        self.wfile.write(body)

    def _redirect(self, location: str, extra_headers: list[tuple[str, str]] | None = None) -> None:
        self.send_response(302)
        self.send_header("Location", location)
        for header_name, header_value in extra_headers or []:
            self.send_header(header_name, header_value)
        self.end_headers()

    def _read_json_body(self) -> dict:
        try:
            content_length = int(self.headers.get("Content-Length", "0"))
        except ValueError as exc:
            raise ValueError("Некорректная длина запроса") from exc

        try:
            raw_body = self.rfile.read(content_length) if content_length > 0 else b"{}"
            return json.loads(raw_body.decode("utf-8"))
        except (UnicodeDecodeError, json.JSONDecodeError) as exc:
            raise ValueError("Некорректный JSON") from exc

    def _build_cookie_header(self, token: str, max_age: int) -> str:
        cookie = cookies.SimpleCookie()
        cookie[SESSION_COOKIE_NAME] = token
        cookie[SESSION_COOKIE_NAME]["path"] = "/"
        cookie[SESSION_COOKIE_NAME]["httponly"] = True
        cookie[SESSION_COOKIE_NAME]["samesite"] = "Lax"
        cookie[SESSION_COOKIE_NAME]["max-age"] = str(max_age)
        return cookie.output(header="").strip()

    def _get_session_token(self) -> str | None:
        raw_cookie = self.headers.get("Cookie")
        if not raw_cookie:
            return None

        jar = cookies.SimpleCookie()
        jar.load(raw_cookie)
        morsel = jar.get(SESSION_COOKIE_NAME)
        return morsel.value if morsel else None

    def _get_current_user(self) -> dict | None:
        return get_user_by_session(self.db_path, self._get_session_token())

    def do_GET(self) -> None:
        parsed_url = urlparse(self.path)
        path = parsed_url.path
        user = self._get_current_user()

        if path == "/":
            if user:
                self._redirect("/admin" if user.get("role") == "admin" else "/dashboard")
                return
            self._send_html("login.html")
            return

        if path == "/dashboard":
            if not user:
                self._redirect("/")
                return
            if user.get("role") == "admin":
                self._redirect("/admin")
                return
            self._send_html("dashboard.html")
            return

        if path == "/admin":
            if not user:
                self._redirect("/?mode=admin")
                return
            if user.get("role") != "admin":
                self._redirect("/dashboard")
                return
            self._send_html("admin.html")
            return

        if path == "/logout":
            delete_session(self.db_path, self._get_session_token())
            self._redirect("/", [("Set-Cookie", self._build_cookie_header("", 0))])
            return

        if path == "/api/me":
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            self._send_json(200, {"ok": True, "user": user})
            return

        if path == "/api/admin/users":
            if not user:
                self._send_json(401, {"ok": False, "error": "????????? ????"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "?????? ?????? ??? ??????????????"})
                return
            payload = list_school_users(self.db_path)
            students_payload: list[dict[str, object]] = []
            for item in payload.get("students", []):
                enriched = dict(item)
                storage_key = str(enriched.get("storage_key") or "").strip()
                if storage_key:
                    ensure_user_storage(self.data_root, storage_key)
                    grades_payload = load_student_grades(self.data_root, storage_key)
                    profile = grades_payload.get("points_profile") or build_points_profile(grades_payload.get("points"))
                    enriched["points_score"] = int((grades_payload.get("points") or {}).get("score") or 0)
                    enriched["points_title"] = str(profile.get("title") or "")
                students_payload.append(enriched)
            payload["students"] = students_payload
            self._send_json(200, {"ok": True, "data": payload})
            return

        if path == "/api/homework":
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            ensure_user_storage(self.data_root, user["storage_key"])
            if user.get("role") == "student":
                state = sync_student_homework_rating(self.data_root, user["storage_key"], date.today().isoformat())
            else:
                from user_store import load_state
                state = load_state(self.data_root, user["storage_key"])
            self._send_json(200, {"ok": True, "role": user.get("role"), "homework": state.get("homework", [])})
            return

        if path == "/api/notes":
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            ensure_user_storage(self.data_root, user["storage_key"])
            if user.get("role") == "student":
                load_student_grades(self.data_root, user["storage_key"])
            state = __import__("user_store").load_state(self.data_root, user["storage_key"])
            self._send_json(200, {"ok": True, "notes": state.get("notes", [])})
            return

        if path == "/api/performance":
            if not user:
                self._send_json(401, {"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚"})
                return
            ensure_user_storage(self.data_root, user["storage_key"])
            performance_payload = (
                __import__("user_store").default_teacher_performance()
                if user.get("role") == "teacher"
                else load_student_performance(self.data_root, user["storage_key"])
            )
            self._send_json(200, {"ok": True, "role": user.get("role"), "performance": performance_payload})
            return

        if path == "/api/grades":
            if not user:
                self._send_json(401, {"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚"})
                return
            ensure_user_storage(self.data_root, user["storage_key"])
            requested_quarter_id = parse_qs(parsed_url.query).get("quarter", [""])[0].strip() or None
            if user.get("role") == "teacher":
                grades_payload = _build_teacher_grades_payload(self.db_path, self.data_root, user, requested_quarter_id)
            else:
                sync_student_homework_rating(self.data_root, user["storage_key"], date.today().isoformat())
                grades_payload = load_student_grades(self.data_root, user["storage_key"])
            self._send_json(200, {"ok": True, "role": user.get("role"), "grades": grades_payload})
            return

        if path == "/api/chat/history":
            if not user:
                self._send_json(401, {"ok": False, "error": "РўСЂРµР±СѓРµС‚СЃСЏ РІС…РѕРґ"})
                return
            self._send_json(200, {"ok": True, "items": list_chat_sessions(self.data_root, user["storage_key"])})
            return

        if path == "/api/chat/thread":
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            chat_id = parse_qs(parsed_url.query).get("chat_id", [""])[0].strip()
            if not chat_id:
                self._send_json(400, {"ok": False, "error": "Не указан chat_id"})
                return
            session = get_chat_session(self.data_root, user["storage_key"], chat_id)
            if session is None:
                self._send_json(404, {"ok": False, "error": "Чат не найден"})
                return
            self._send_json(200, {"ok": True, "session": session})
            return

        if path == "/api/homework/submission":
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "teacher":
                self._send_json(403, {"ok": False, "error": "Недостаточно прав"})
                return

            query = parse_qs(parsed_url.query)
            student_username = str(query.get("student", [""])[0] or "").strip()
            homework_id = str(query.get("homework_id", [""])[0] or "").strip()
            if not student_username or not homework_id:
                self._send_json(400, {"ok": False, "error": "Не указаны параметры загрузки"})
                return

            managed_classes = get_teacher_managed_classes(self.db_path, str(user.get("username") or ""))
            students = list_students_for_classes(self.db_path, managed_classes)
            student = next((item for item in students if str(item.get("username") or "").strip() == student_username), None)
            if student is None:
                self._send_json(404, {"ok": False, "error": "Ученик не найден"})
                return

            storage_key = str(student.get("storage_key") or "").strip()
            submission = get_student_homework_submission(self.data_root, storage_key, homework_id)
            if submission is None:
                self._send_json(404, {"ok": False, "error": "Файл решения не найден"})
                return

            self._send_file(str(submission["file_path"]), str(submission["file_name"]))
            return

        if path == "/api/notifications":
            if not user:
                self._send_json(401, {"ok": False, "error": "????????? ????"})
                return
            if user.get("role") != "student":
                self._send_json(200, {"ok": True, "notifications": []})
                return
            notifications = build_student_ai_notifications(
                self.data_root,
                user["storage_key"],
                date.today().isoformat(),
            )
            self._send_json(200, {"ok": True, "notifications": notifications})
            return

        if path == "/api/archive":
            if not user:
                self._send_json(401, {"ok": False, "error": "??????? ??????? ? ???????"})
                return
            if user.get("role") == "teacher":
                query = parse_qs(parsed_url.query)
                archive = _build_teacher_archive_payload(
                    self.db_path,
                    self.data_root,
                    user,
                    selected_subject=str(query.get("subject", [""])[0] or "").strip() or None,
                    selected_date=str(query.get("date", [""])[0] or "").strip() or None,
                    selected_class=str(query.get("class_name", [""])[0] or "").strip() or None,
                )
                self._send_json(200, {"ok": True, "archive": archive, "mode": "teacher"})
                return
            archive = archive_overdue_notes(self.data_root, user["storage_key"], date.today().isoformat())
            self._send_json(200, {"ok": True, "archive": archive, "mode": "student"})
            return

        super().do_GET()

    def do_POST(self) -> None:
        path = urlparse(self.path).path

        if path == "/api/login":
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            username = str(payload.get("username", "")).strip()
            password = str(payload.get("password", "")).strip()

            if not username or not password:
                self._send_json(400, {"ok": False, "error": "Введите логин и пароль"})
                return

            user = authenticate_user(self.db_path, username, password)
            if user is None:
                self._send_json(401, {"ok": False, "error": "Неверный логин или пароль"})
                return

            session_token = create_session(self.db_path, user["id"])
            headers = [("Set-Cookie", self._build_cookie_header(session_token, 7 * 24 * 60 * 60))]
            self._send_json(
                200,
                {
                    "ok": True,
                    "user": {
                        "username": user["username"],
                        "role": user["role"],
                        "display_name": user["display_name"],
                    },
                    "redirect": "/admin" if user["role"] == "admin" else "/dashboard",
                },
                headers,
            )
            return

        if path == "/api/homework/toggle":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            hw_id = str(payload.get("id", "")).strip()
            if user.get("role") == "student":
                toggled, info = toggle_student_homework_completion(self.data_root, user["storage_key"], hw_id)
                if not toggled:
                    self._send_json(404, {"ok": False, "error": "Не найдено домашнее задание"})
                    return
                self._send_json(200, {"ok": True, "done": info.get("done"), "points": info.get("points"), "points_profile": info.get("points_profile"), "streak": info.get("streak")})
                return
            state = load_state(self.data_root, user["storage_key"])
            for hw in state.get("homework", []):
                if hw.get("id") == hw_id:
                    hw["done"] = not hw.get("done", False)
                    break
            save_state(self.data_root, user["storage_key"], state)
            self._send_json(200, {"ok": True})
            return

        if path == "/api/homework/delete":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "teacher":
                self._send_json(403, {"ok": False, "error": "Недостаточно прав"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            hw_id = str(payload.get("id", "")).strip()
            if not hw_id:
                self._send_json(400, {"ok": False, "error": "Не указан id"})
                return

            teacher_state = load_state(self.data_root, user["storage_key"])
            target_homework = next(
                (
                    item for item in teacher_state.get("homework", [])
                    if str(item.get("id") or "").strip() == hw_id and str(item.get("source") or "").strip() == "teacher"
                ),
                None,
            )
            if target_homework is None:
                self._send_json(404, {"ok": False, "error": "Домашнее задание не найдено"})
                return

            class_name = str(target_homework.get("class_name") or "").strip()
            students = list_students_for_classes(self.db_path, [class_name] if class_name else [])
            deleted = delete_teacher_homework_assignment(self.data_root, user["storage_key"], hw_id, students)
            if deleted is None:
                self._send_json(404, {"ok": False, "error": "Домашнее задание не найдено"})
                return

            self._send_json(200, {"ok": True, "deleted": deleted})
            return

        if path == "/api/homework/submit":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "????????? ????"})
                return
            if user.get("role") != "student":
                self._send_json(403, {"ok": False, "error": "???????????? ????"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            hw_id = str(payload.get("id", "")).strip()
            submission_text = str(payload.get("submission_text", "")).strip()
            file_name = str(payload.get("file_name", "")).strip()
            raw_content = str(payload.get("content", "")).strip()
            if not hw_id:
                self._send_json(400, {"ok": False, "error": "?? ?????? id"})
                return
            if not file_name or not raw_content:
                self._send_json(400, {"ok": False, "error": "?? ?????? ???? ???????"})
                return

            try:
                file_bytes = base64.b64decode(raw_content, validate=True)
            except (ValueError, TypeError):
                self._send_json(400, {"ok": False, "error": "?? ??????? ?????????? ????"})
                return

            submitted, info = submit_student_homework(
                self.data_root,
                user["storage_key"],
                hw_id,
                submission_text=submission_text,
                file_name=file_name,
                file_bytes=file_bytes,
            )
            if not submitted:
                self._send_json(404, {"ok": False, "error": "???????? ??????? ?? ???????"})
                return

            self._send_json(200, {"ok": True, **(info or {})})
            return

        if path == "/api/notes/delete":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            note_id = str(payload.get("id", "")).strip()
            if not note_id:
                self._send_json(400, {"ok": False, "error": "Не указан id"})
                return
            state = load_state(self.data_root, user["storage_key"])
            state["notes"] = [n for n in state.get("notes", []) if n.get("id") != note_id]
            save_state(self.data_root, user["storage_key"], state)
            self._send_json(200, {"ok": True})
            return

        if path == "/api/notes/complete":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            note_id = str(payload.get("id", "")).strip()
            if not note_id:
                self._send_json(400, {"ok": False, "error": "РќРµ СѓРєР°Р·Р°РЅ id"})
                return
            archived_note = complete_note(self.data_root, user["storage_key"], note_id)
            if archived_note is None:
                self._send_json(404, {"ok": False, "error": "РќР°РїРѕРјРёРЅР°РЅРёРµ РЅРµ РЅР°Р№РґРµРЅРѕ"})
                return
            self._send_json(200, {"ok": True, "item": archived_note})
            return

        if path == "/api/notes/subtasks":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "????????? ????"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            note_id = str(payload.get("id", "")).strip()
            if not note_id:
                self._send_json(400, {"ok": False, "error": "?? ?????? id"})
                return

            state = load_state(self.data_root, user["storage_key"])
            note = next((item for item in state.get("notes", []) if str(item.get("id") or "").strip() == note_id), None)
            if note is None:
                self._send_json(404, {"ok": False, "error": "??????? ?? ???????"})
                return

            result = note.get("result") if isinstance(note.get("result"), dict) else {}
            note_text = str(result.get("task") or note.get("source_text") or "").strip()
            if not note_text:
                self._send_json(400, {"ok": False, "error": "? ??????? ??? ?????? ??? ?????????"})
                return

            try:
                subtasks = generate_note_subtasks(note_text)
            except Exception:
                self._send_json(500, {"ok": False, "error": "?? ??????? ??????? ??????? ?? ?????"})
                return

            updated_note = save_note_subtasks(self.data_root, user["storage_key"], note_id, subtasks)
            if updated_note is None:
                self._send_json(404, {"ok": False, "error": "??????? ?? ???????"})
                return

            self._send_json(200, {"ok": True, "note": updated_note, "subtasks": subtasks})
            return

        if path == "/api/notes/subtasks/toggle":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "????????? ????"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            note_id = str(payload.get("id", "")).strip()
            try:
                subtask_index = int(payload.get("subtask_index"))
            except (TypeError, ValueError):
                self._send_json(400, {"ok": False, "error": "???????????? ?????? ?????????"})
                return

            if not note_id:
                self._send_json(400, {"ok": False, "error": "?? ?????? id"})
                return

            updated_note = toggle_note_subtask(self.data_root, user["storage_key"], note_id, subtask_index)
            if updated_note is None:
                self._send_json(404, {"ok": False, "error": "????????? ?? ???????"})
                return

            self._send_json(200, {"ok": True, "note": updated_note})
            return

        if path == "/api/archive/item":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            item_id = str(payload.get("id", "")).strip()
            section = str(payload.get("section", "")).strip()
            if not item_id or not section:
                self._send_json(400, {"ok": False, "error": "РќРµ РїРµСЂРµРґР°РЅС‹ РґР°РЅРЅС‹Рµ РєР°СЂС‚РѕС‡РєРё"})
                return
            archive = move_archive_item_to_trash(self.data_root, user["storage_key"], section, item_id)
            self._send_json(200, {"ok": True, "archive": archive})
            return

        if path == "/api/archive/trash/clear":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "РЎРЅР°С‡Р°Р»Р° РІРѕР№РґРёС‚Рµ РІ Р°РєРєР°СѓРЅС‚"})
                return
            archive = clear_archive_trash(self.data_root, user["storage_key"])
            self._send_json(200, {"ok": True, "archive": archive})
            return

        if path == "/api/admin/students":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "Доступ только для администратора"})
                return
            try:
                payload = self._read_json_body()
                student = create_student_user(
                    self.db_path,
                    str(payload.get("display_name", "")),
                    str(payload.get("class_name", "")),
                    str(payload.get("username", "")),
                    str(payload.get("password", "")),
                )
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            except Exception:
                self._send_json(500, {"ok": False, "error": "Не удалось создать ученика"})
                return
            self._send_json(200, {"ok": True, "student": student, "data": list_school_users(self.db_path)})
            return

        if path == "/api/admin/students/import":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "Доступ только для администратора"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            file_name = str(payload.get("file_name", "")).strip()
            content = str(payload.get("content", "")).strip()
            if not file_name or not content:
                self._send_json(400, {"ok": False, "error": "Не выбран файл для импорта"})
                return

            try:
                file_bytes = base64.b64decode(content, validate=True)
                students = _extract_students_from_import(file_name, file_bytes)
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            except Exception:
                self._send_json(500, {"ok": False, "error": "Не удалось обработать файл"})
                return

            if not students:
                self._send_json(400, {"ok": False, "error": "В таблице не найдено строк для импорта"})
                return

            created = 0
            for student in students:
                try:
                    create_student_user(
                        self.db_path,
                        student.get("display_name", ""),
                        student.get("class_name", ""),
                        student.get("username", ""),
                        student.get("password", ""),
                    )
                    created += 1
                except ValueError:
                    continue

            self._send_json(
                200,
                {
                    "ok": True,
                    "imported": created,
                    "total": len(students),
                    "data": list_school_users(self.db_path),
                },
            )
            return

        if path == "/api/admin/teachers":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "Доступ только для администратора"})
                return
            try:
                payload = self._read_json_body()
                teacher = create_teacher_user(
                    self.db_path,
                    str(payload.get("display_name", "")),
                    str(payload.get("managed_classes", "")),
                    str(payload.get("username", "")),
                    str(payload.get("password", "")),
                )
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            except Exception:
                self._send_json(500, {"ok": False, "error": "Не удалось создать учителя"})
                return
            self._send_json(200, {"ok": True, "teacher": teacher, "data": list_school_users(self.db_path)})
            return

        if path == "/api/admin/students/delete":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "Доступ только для администратора"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            try:
                user_id = int(payload.get("id"))
            except (TypeError, ValueError):
                self._send_json(400, {"ok": False, "error": "Некорректный id ученика"})
                return
            if not delete_student_user(self.db_path, user_id):
                self._send_json(404, {"ok": False, "error": "Ученик не найден"})
                return
            self._send_json(200, {"ok": True, "data": list_school_users(self.db_path)})
            return

        if path == "/api/admin/teachers/delete":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Требуется вход"})
                return
            if user.get("role") != "admin":
                self._send_json(403, {"ok": False, "error": "Доступ только для администратора"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            try:
                user_id = int(payload.get("id"))
            except (TypeError, ValueError):
                self._send_json(400, {"ok": False, "error": "Некорректный id учителя"})
                return
            if not delete_teacher_user(self.db_path, user_id):
                self._send_json(404, {"ok": False, "error": "Учитель не найден"})
                return
            self._send_json(200, {"ok": True, "data": list_school_users(self.db_path)})
            return

        if path == "/api/process":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Сначала войдите в аккаунт"})
                return

            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            text = str(payload.get("text", "")).strip()
            if not text:
                self._send_json(400, {"ok": False, "error": "Введите текст для обработки"})
                return

            try:
                result = smart_processor(text)
                saved_entry = save_planner_result(self.data_root, user["storage_key"], text, result)

                # Если это ДЗ (категория 3) — создаём задание для учеников
                if result.get("category") == 3:
                    if user.get("role") != "teacher":
                        self._send_json(400, {"ok": False, "error": "Создавать домашние задания может только учитель"})
                        return
                    class_name = str(result.get("class_name") or "").strip()
                    task = str(result.get("task") or "").strip()
                    date = str(result.get("date") or "").strip()
                    subject = str(result.get("subject") or "").strip()
                    if not class_name or not task or not date:
                        self._send_json(400, {"ok": False, "error": "Для создания ДЗ нужны класс, дата и текст задания"})
                        return
                    managed_classes = get_teacher_managed_classes(self.db_path, user.get("username"))
                    if managed_classes and class_name not in managed_classes:
                        self._send_json(400, {"ok": False, "error": f"Класс {class_name} не закреплён за этим учителем"})
                        return
                    students = list_students_for_classes(self.db_path, [class_name])
                    assignment = create_teacher_homework_assignment(
                        self.data_root,
                        user["storage_key"],
                        students,
                        class_name=class_name,
                        subject=subject,
                        task=task,
                        date=date,
                        points_value=int(result.get("points_value") or 1),
                        teacher_name=user.get("display_name"),
                    )
                    self._send_json(200, {
                        "ok": True,
                        "result": result,
                        "homework_saved": True,
                        "homework": {
                            "class_name": class_name,
                            "subject": subject,
                            "task": task,
                            "date": date,
                            "created_students": assignment["created_students"],
                        },
                    })
                    return

                if result.get("category") == 1:
                    self._send_json(
                        200,
                        {
                            "ok": True,
                            "result": result,
                            "reminder_saved": True,
                            "reminder_entry": saved_entry,
                        },
                    )
                    return

                if result.get("category") == 4:
                    self._send_json(200, {"ok": True, "result": result})
                    return

                self._send_json(200, {"ok": True, "result": result})
            except Exception:
                self._send_json(500, {"ok": False, "error": "Не удалось обработать запрос"})
            return

        if path == "/api/chat/send":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Сначала войдите в аккаунт"})
                return

            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            message = str(payload.get("message", "")).strip()
            chat_id = str(payload.get("chat_id", "")).strip()
            chat_id = None if not chat_id or chat_id == "None" else chat_id
            if not message:
                self._send_json(400, {"ok": False, "error": "Введите сообщение"})
                return

            # --- Фильтрация через smart_processor ---
            # Если пользователь хочет создать заметку/напоминание — сохраняем и отвечаем подтверждением,
            # иначе — идём в обычный chat_reply.
            try:
                parsed = smart_processor(message)
            except Exception:
                parsed = {"category": 2}

            if parsed.get("category") == 3:
                if user.get("role") != "teacher":
                    self._send_json(400, {"ok": False, "error": "Создавать домашние задания через чат может только учитель"})
                    return

                class_name = str(parsed.get("class_name") or "").strip()
                task = str(parsed.get("task") or "").strip()
                date = str(parsed.get("date") or "").strip()
                subject = str(parsed.get("subject") or "").strip()

                if not class_name or not task or not date:
                    self._send_json(400, {"ok": False, "error": "Для создания ДЗ нужны класс, дата и текст задания"})
                    return

                managed_classes = get_teacher_managed_classes(self.db_path, user.get("username"))
                if managed_classes and class_name not in managed_classes:
                    self._send_json(400, {"ok": False, "error": f"Класс {class_name} не закреплен за этим учителем"})
                    return

                students = list_students_for_classes(self.db_path, [class_name])
                assignment = create_teacher_homework_assignment(
                    self.data_root,
                    user["storage_key"],
                    students,
                    class_name=class_name,
                    subject=subject,
                    task=task,
                    date=date,
                    points_value=int(parsed.get("points_value") or 1),
                    teacher_name=user.get("display_name"),
                )
                assistant_message = (
                    f"✅ Домашнее задание создано для {class_name}: «{task}» на {date}. "
                    f"Ученикам добавлено: {assignment['created_students']}."
                )
                session = save_chat_exchange(
                    self.data_root,
                    user["storage_key"],
                    user_message=message,
                    assistant_message=assistant_message,
                    chat_id=chat_id,
                )
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "chat_id": session["chat_id"],
                        "messages": session["messages"],
                        "history": list_chat_sessions(self.data_root, user["storage_key"]),
                        "homework_saved": True,
                        "homework": {
                            "class_name": class_name,
                            "subject": subject,
                            "task": task,
                            "date": date,
                            "created_students": assignment["created_students"],
                        },
                    },
                )
                return

            if parsed.get("category") == 4:
                try:
                    analysis_context = _build_analysis_context(self.data_root, user["storage_key"], parsed)
                    assistant_message = generate_analysis_reply(message, analysis_context)
                except Exception:
                    self._send_json(500, {"ok": False, "error": "Не удалось выполнить анализ данных"})
                    return

                session = save_chat_exchange(
                    self.data_root,
                    user["storage_key"],
                    user_message=message,
                    assistant_message=assistant_message,
                    chat_id=chat_id,
                )
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "chat_id": session["chat_id"],
                        "messages": session["messages"],
                        "history": list_chat_sessions(self.data_root, user["storage_key"]),
                        "analysis": {
                            "scope": parsed.get("scope", []),
                            "plan_date": parsed.get("plan_date"),
                            "focus": parsed.get("focus"),
                        },
                    },
                )
                return

            if parsed.get("category") == 1:
                # Сохраняем заметку в планировщик
                try:
                    save_planner_result(self.data_root, user["storage_key"], message, parsed)
                except Exception:
                    pass

                task = parsed.get("task") or message
                date = parsed.get("date")
                if date:
                    assistant_message = f"✅ Напоминание сохранено: «{task}» на {date}."
                else:
                    assistant_message = f"✅ Напоминание сохранено: «{task}»."

                session = save_chat_exchange(
                    self.data_root,
                    user["storage_key"],
                    user_message=message,
                    assistant_message=assistant_message,
                    chat_id=chat_id,
                )
                self._send_json(
                    200,
                    {
                        "ok": True,
                        "chat_id": session["chat_id"],
                        "messages": session["messages"],
                        "history": list_chat_sessions(self.data_root, user["storage_key"]),
                        "reminder_saved": True,
                        "reminder": {"task": task, "date": date},
                    },
                )
                return
            # --- конец фильтрации ---

            existing_session = get_chat_session(self.data_root, user["storage_key"], chat_id) if chat_id else None
            conversation = []
            if existing_session:
                conversation.extend(
                    {"role": item["role"], "content": item["content"]}
                    for item in existing_session.get("messages", [])
                )
            conversation.append({"role": "user", "content": message})

            try:
                assistant_message = chat_reply(conversation)
            except Exception:
                self._send_json(500, {"ok": False, "error": "Не удалось получить ответ от модели"})
                return

            session = save_chat_exchange(
                self.data_root,
                user["storage_key"],
                user_message=message,
                assistant_message=assistant_message,
                chat_id=chat_id,
            )
            self._send_json(
                200,
                {
                    "ok": True,
                    "chat_id": session["chat_id"],
                    "messages": session["messages"],
                    "history": list_chat_sessions(self.data_root, user["storage_key"]),
                },
            )
            return

        if path == "/api/tts":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Не авторизован"})
                return
            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return
            text = str(payload.get("text", "")).strip()
            voice = str(payload.get("voice", "male")).strip()
            if not text:
                self._send_json(400, {"ok": False, "error": "Нет текста"})
                return
            try:
                from tts_engine import synthesize_to_bytes
                wav = synthesize_to_bytes(text, gender=voice)
                self.send_response(200)
                self.send_header("Content-Type", "audio/wav")
                self.send_header("Content-Length", str(len(wav)))
                self.end_headers()
                self.wfile.write(wav)
            except Exception as exc:
                self._send_json(500, {"ok": False, "error": str(exc)})
            return

        if path == "/api/settings/model":
            user = self._get_current_user()
            if not user:
                self._send_json(401, {"ok": False, "error": "Сначала войдите в аккаунт"})
                return

            try:
                payload = self._read_json_body()
            except ValueError as error:
                self._send_json(400, {"ok": False, "error": str(error)})
                return

            allowed_models = {"qwen2.5:3b", "mistral:7b-instruct-v0.3-q4_0"}
            model = str(payload.get("model", "")).strip()
            if model not in allowed_models:
                self._send_json(400, {"ok": False, "error": f"Недопустимая модель: {model}"})
                return

            import planner
            planner.MODEL = model
            self._send_json(200, {"ok": True, "model": model})
            return

        self._send_json(404, {"ok": False, "error": "Маршрут не найден"})


def main() -> None:
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    project_dir = os.path.dirname(os.path.abspath(__file__))
    templates_dir = os.path.join(project_dir, "templates")
    base_dir = templates_dir if os.path.isdir(templates_dir) else project_dir
    db_path = os.path.join(project_dir, DB_FILENAME)
    data_root = os.path.join(project_dir, "user_data")

    init_database(db_path)
    os.makedirs(data_root, exist_ok=True)

    AppHandler.db_path = db_path
    AppHandler.data_root = data_root
    handler = partial(AppHandler, directory=base_dir)
    server = ThreadingHTTPServer((host, port), handler)
    print(f"Server started: http://{host}:{port}")
    print(f"Serving directory: {base_dir}")
    print(f"SQLite DB: {db_path}")
    print(f"User data root: {data_root}")
    print("POST API: /api/login, /api/process, /api/chat/send")
    print("Press Ctrl+C to stop")

    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nServer stopped")
    finally:
        server.server_close()


if __name__ == "__main__":
    main()
